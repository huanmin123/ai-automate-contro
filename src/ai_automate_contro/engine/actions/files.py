from __future__ import annotations

import csv
from copy import copy
from datetime import date, datetime, time, timedelta
from decimal import Decimal
import json
import re
from typing import Any


EXCEL_A1_RE = re.compile(r"^[A-Za-z]{1,3}[1-9][0-9]*(?::[A-Za-z]{1,3}[1-9][0-9]*)?$")


def write_json_file(executor: Any, raw_path: str, value: Any, *, category: str = "json", indent: int = 2) -> None:
    output_path = executor._resolve_output_path(raw_path, category=category)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with output_path.open("w", encoding="utf-8") as file:
        json.dump(value, file, ensure_ascii=False, indent=indent)
    executor.state.logger.log("info", "json written", path=str(output_path))


def write_text_file(executor: Any, raw_path: str, content: Any, *, append: bool) -> None:
    output_path = executor._resolve_output_path(raw_path, category="text")
    output_path.parent.mkdir(parents=True, exist_ok=True)
    mode = "a" if append else "w"
    with output_path.open(mode, encoding="utf-8") as file:
        file.write(_coerce_text_content(content))
    log_message = "text appended" if append else "text written"
    executor.state.logger.log("info", log_message, path=str(output_path))


def write_csv_file(executor: Any, raw_path: str, rows: list[Any], headers: list[str] | None = None) -> None:
    output_path = executor._resolve_output_path(raw_path, category="csv")
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with output_path.open("w", encoding="utf-8-sig", newline="") as file:
        if headers:
            writer = csv.writer(file)
            writer.writerow(headers)
            for row in rows:
                if isinstance(row, dict):
                    writer.writerow([row.get(header, "") for header in headers])
                else:
                    writer.writerow(row)
        elif not rows:
            file.write("")
        elif isinstance(rows[0], dict):
            fieldnames = list(rows[0].keys())
            writer = csv.DictWriter(file, fieldnames=fieldnames)
            writer.writeheader()
            writer.writerows(rows)
        else:
            writer = csv.writer(file)
            writer.writerows(rows)
    executor.state.logger.log("info", "csv written", path=str(output_path))


def write_excel_file(executor: Any, step: dict[str, Any]) -> None:
    excel = _load_openpyxl()
    output_path = executor._resolve_output_path(step["path"], category="excel")
    output_path.parent.mkdir(parents=True, exist_ok=True)

    template_path = step.get("template_path")
    if template_path:
        source_path = executor._resolve_path(template_path)
        workbook = excel["load_workbook"](source_path, keep_vba=source_path.suffix.lower() == ".xlsm")
    else:
        workbook = excel["Workbook"]()

    sheet_steps = _excel_write_sheet_steps(step)
    for sheet_step in sheet_steps:
        _write_excel_sheet(
            workbook,
            sheet_step,
            template=bool(template_path),
            get_column_letter=excel["get_column_letter"],
            range_boundaries=excel["range_boundaries"],
        )

    workbook.save(output_path)
    executor.state.logger.log("info", "excel written", path=str(output_path))


def read_file(executor: Any, step: dict[str, Any]) -> Any:
    file_type = step["type"]
    path = executor._resolve_path(step["path"])
    if file_type == "json":
        with path.open("r", encoding="utf-8") as file:
            return json.load(file)
    if file_type == "text":
        with path.open("r", encoding="utf-8") as file:
            content = file.read()
        if step.get("split_lines", False):
            return [line.strip() for line in content.splitlines() if line.strip()]
        return content
    if file_type == "csv":
        with path.open("r", encoding="utf-8-sig", newline="") as file:
            reader = csv.DictReader(file)
            return list(reader)
    if file_type == "excel":
        value, _meta = read_excel_file(path, step)
        return value
    if file_type == "storage_state":
        return str(path)
    raise ValueError(f"Unsupported read type: {file_type}")


def read_excel_file(path: Any, step: dict[str, Any]) -> tuple[Any, dict[str, Any]]:
    excel = _load_openpyxl()
    formula_mode = str(step.get("formula_mode", "cached"))
    workbook = excel["load_workbook"](path, read_only=True, data_only=formula_mode != "formula")
    try:
        if "sheets" in step:
            return _read_excel_sheets(path, step, excel, workbook)
        return _read_excel_workbook(path, step, excel, workbook)
    finally:
        workbook.close()


def _read_excel_sheets(path: Any, step: dict[str, Any], excel: dict[str, Any], workbook: Any) -> tuple[dict[str, Any], dict[str, Any]]:
    sheet_steps = _excel_read_sheet_steps(step)
    values: dict[str, Any] = {}
    sheet_meta: dict[str, Any] = {}
    for sheet_step in sheet_steps:
        value, meta = _read_excel_workbook(path, sheet_step, excel, workbook)
        sheet_name = str(meta.get("sheet") or sheet_step.get("sheet") or "")
        if not sheet_name:
            raise ValueError("read.type=excel.sheets 每一项必须能解析出 sheet 名称。")
        value_name = str(sheet_step.get("name") or sheet_name)
        if not value_name:
            raise ValueError("read.type=excel.sheets 每一项的 name 必须是非空字符串。")
        if value_name in values:
            raise ValueError(f"read.type=excel.sheets 不能重复保存同名结果：{value_name}")
        values[value_name] = value
        sheet_meta[value_name] = meta
    meta = {
        "type": "excel",
        "path": str(path),
        "sheets": list(workbook.sheetnames),
        "selected_sheets": [str(meta.get("sheet") or key) for key, meta in sheet_meta.items()],
        "value_names": list(values.keys()),
        "sheet_meta": sheet_meta,
        "row_count": sum(int(meta.get("row_count", 0) or 0) for meta in sheet_meta.values()),
    }
    return values, meta


def _excel_read_sheet_steps(step: dict[str, Any]) -> list[dict[str, Any]]:
    raw_sheets = step.get("sheets")
    if not isinstance(raw_sheets, list) or not raw_sheets:
        raise ValueError("read.type=excel.sheets 必须是非空数组。")
    sheet_steps: list[dict[str, Any]] = []
    for index, raw_sheet in enumerate(raw_sheets):
        merged = {
            key: value
            for key, value in step.items()
            if key not in {"sheets", "sheet", "output"}
        }
        if isinstance(raw_sheet, dict):
            merged.update(raw_sheet)
        elif isinstance(raw_sheet, (str, int)):
            merged["sheet"] = raw_sheet
        else:
            raise ValueError(f"read.type=excel.sheets[{index}] 必须是 sheet 名称、索引或读取配置对象。")
        if "sheet" not in merged:
            raise ValueError(f"read.type=excel.sheets[{index}] 缺少 sheet。")
        sheet_steps.append(merged)
    return sheet_steps


def _read_excel_workbook(path: Any, step: dict[str, Any], excel: dict[str, Any], workbook: Any) -> tuple[Any, dict[str, Any]]:
    date_format = str(step.get("date_format", "iso"))
    worksheet = _select_excel_worksheet(workbook, step.get("sheet"))
    min_col, min_row, max_col, max_row = _excel_bounds(worksheet, step.get("range"), excel["range_boundaries"])
    mode = str(step.get("mode", "records"))
    max_rows = int(step.get("max_rows", 10000))
    max_cells = int(step.get("max_cells", 500000))
    preview_rows = _optional_positive_int(step.get("preview_rows"), field_name="preview_rows")
    offset_rows = _optional_nonnegative_int(step.get("offset_rows"), field_name="offset_rows") or 0
    limit_rows = _optional_positive_int(step.get("limit_rows"), field_name="limit_rows")
    skip_blank_rows = bool(step.get("skip_blank_rows", True))

    if max_row < min_row or max_col < min_col:
        value: Any = [] if mode in {"records", "matrix"} else {}
        meta = _excel_meta(path, workbook, worksheet, "", [], value)
        return value, meta

    if mode == "matrix":
        original_max_row = max_row
        data_min_row, data_max_row, truncated = _excel_row_window(
            min_row,
            max_row,
            offset_rows=offset_rows,
            limit_rows=limit_rows,
            preview_rows=preview_rows,
        )
        _ensure_max_cells(min_col, data_min_row, max_col, data_max_row, max_cells)
        rows = [
            [_json_safe_excel_value(cell, date_format=date_format) for cell in row]
            for row in worksheet.iter_rows(
                min_row=data_min_row,
                max_row=data_max_row,
                min_col=min_col,
                max_col=max_col,
                values_only=True,
            )
            if not skip_blank_rows or not _excel_row_is_blank(row)
        ] if data_min_row <= data_max_row else []
        _ensure_max_rows(rows, max_rows)
        meta = _excel_meta(
            path,
            workbook,
            worksheet,
            _excel_range_text(min_col, data_min_row, max_col, data_max_row, excel["get_column_letter"]),
            [],
            rows,
        )
        _apply_excel_read_window_meta(
            meta,
            max_cells=max_cells,
            offset_rows=offset_rows,
            limit_rows=limit_rows,
            preview_rows=preview_rows,
            truncated=truncated or original_max_row > data_max_row,
        )
        return rows, meta

    if mode == "cells":
        original_max_row = max_row
        data_min_row, data_max_row, truncated = _excel_row_window(
            min_row,
            max_row,
            offset_rows=offset_rows,
            limit_rows=limit_rows,
            preview_rows=preview_rows,
        )
        _ensure_max_cells(min_col, data_min_row, max_col, data_max_row, max_cells)
        cells: dict[str, Any] = {}
        if data_min_row <= data_max_row:
            for row in worksheet.iter_rows(min_row=data_min_row, max_row=data_max_row, min_col=min_col, max_col=max_col):
                for cell in row:
                    value = _json_safe_excel_value(cell.value, date_format=date_format)
                    if value is not None and value != "":
                        cells[cell.coordinate] = value
        meta = _excel_meta(
            path,
            workbook,
            worksheet,
            _excel_range_text(min_col, data_min_row, max_col, data_max_row, excel["get_column_letter"]),
            [],
            cells,
        )
        _apply_excel_read_window_meta(
            meta,
            max_cells=max_cells,
            offset_rows=offset_rows,
            limit_rows=limit_rows,
            preview_rows=preview_rows,
            truncated=truncated or original_max_row > data_max_row,
        )
        return cells, meta

    headers, data_start_row, min_col, max_col = _excel_headers(worksheet, step, min_row, max_row, min_col, max_col)
    original_max_row = max_row
    data_min_row, data_max_row, truncated = _excel_row_window(
        data_start_row,
        max_row,
        offset_rows=offset_rows,
        limit_rows=limit_rows,
        preview_rows=preview_rows,
    )
    _ensure_max_cell_count(
        max_col - min_col + 1,
        (0 if "headers" in step else 1) + max(0, data_max_row - data_min_row + 1),
        max_cells,
    )
    records: list[dict[str, Any]] = []
    if data_min_row <= data_max_row:
        for row in worksheet.iter_rows(
            min_row=data_min_row,
            max_row=data_max_row,
            min_col=min_col,
            max_col=max_col,
            values_only=True,
        ):
            if skip_blank_rows and _excel_row_is_blank(row):
                continue
            records.append({header: _json_safe_excel_value(value, date_format=date_format) for header, value in zip(headers, row)})
            _ensure_max_rows(records, max_rows)

    meta = _excel_meta(
        path,
        workbook,
        worksheet,
        _excel_range_text(min_col, min_row, max_col, data_max_row, excel["get_column_letter"]),
        headers,
        records,
    )
    _apply_excel_read_window_meta(
        meta,
        max_cells=max_cells,
        offset_rows=offset_rows,
        limit_rows=limit_rows,
        preview_rows=preview_rows,
        truncated=truncated or original_max_row > data_max_row,
    )
    return records, meta


def _coerce_text_content(value: Any) -> str:
    if isinstance(value, str):
        return value
    if isinstance(value, (list, tuple)):
        lines = [str(item) for item in value]
        return ("\n".join(lines) + "\n") if lines else ""
    return str(value)


def _load_openpyxl() -> dict[str, Any]:
    try:
        from openpyxl import Workbook, load_workbook
        from openpyxl.worksheet.table import Table, TableStyleInfo
        from openpyxl.utils import get_column_letter, range_boundaries
    except ImportError as error:
        raise RuntimeError("Excel 读写需要安装 openpyxl：pip install openpyxl>=3.1,<4.0") from error
    return {
        "Workbook": Workbook,
        "Table": Table,
        "TableStyleInfo": TableStyleInfo,
        "get_column_letter": get_column_letter,
        "load_workbook": load_workbook,
        "range_boundaries": range_boundaries,
    }


def _select_excel_worksheet(workbook: Any, raw_sheet: Any) -> Any:
    if raw_sheet in (None, ""):
        return workbook.worksheets[0]
    if isinstance(raw_sheet, int):
        try:
            return workbook.worksheets[raw_sheet]
        except IndexError as error:
            raise ValueError(f"Excel sheet 索引超出范围：{raw_sheet}") from error
    sheet_name = str(raw_sheet)
    if sheet_name not in workbook.sheetnames:
        raise ValueError(f"Excel 工作表不存在：{sheet_name}")
    return workbook[sheet_name]


def _excel_bounds(worksheet: Any, raw_range: Any, range_boundaries: Any) -> tuple[int, int, int, int]:
    if raw_range:
        range_text = str(raw_range)
        if not EXCEL_A1_RE.match(range_text):
            raise ValueError(f"Excel range 必须是 A1 风格范围：{range_text}")
        return range_boundaries(range_text)
    try:
        dimension = worksheet.calculate_dimension()
    except ValueError as error:
        if "unsized" not in str(error):
            raise
        dimension = worksheet.calculate_dimension(force=True)
    min_col, min_row, max_col, max_row = range_boundaries(dimension)
    if min_col == max_col and min_row == max_row and not _excel_cell_has_value(worksheet.cell(row=min_row, column=min_col).value):
        return 1, 1, 0, 0
    return min_col, min_row, max_col, max_row


def _excel_cell_has_value(value: Any) -> bool:
    if value is None:
        return False
    if isinstance(value, str) and not value.strip():
        return False
    return True


def _excel_headers(
    worksheet: Any,
    step: dict[str, Any],
    min_row: int,
    max_row: int,
    min_col: int,
    max_col: int,
) -> tuple[list[str], int, int, int]:
    raw_headers = step.get("headers")
    if raw_headers is not None:
        headers = [str(header).strip() for header in raw_headers]
        _validate_excel_headers(headers)
        return headers, min_row, min_col, max_col

    header_row = int(step["header_row"]) if "header_row" in step else _excel_first_nonblank_row(worksheet, min_row, max_row, min_col, max_col)
    values = next(
        worksheet.iter_rows(
            min_row=header_row,
            max_row=header_row,
            min_col=min_col,
            max_col=max_col,
            values_only=True,
        ),
        (),
    )
    header_min_col, header_max_col, values = _excel_header_bounds(values, min_col)
    headers = [str(_json_safe_excel_value(value) or "").strip() for value in values]
    _validate_excel_headers(headers)
    return headers, header_row + 1, header_min_col, header_max_col


def _excel_header_bounds(values: tuple[Any, ...], min_col: int) -> tuple[int, int, tuple[Any, ...]]:
    nonblank_indexes = [index for index, value in enumerate(values) if _excel_cell_has_value(value)]
    if not nonblank_indexes:
        return min_col, min_col + len(values) - 1, values
    start_index = min(nonblank_indexes)
    end_index = max(nonblank_indexes)
    return min_col + start_index, min_col + end_index, tuple(values[start_index : end_index + 1])


def _excel_first_nonblank_row(worksheet: Any, min_row: int, max_row: int, min_col: int, max_col: int) -> int:
    for row_index, row in enumerate(
        worksheet.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col, values_only=True),
        start=min_row,
    ):
        if not _excel_row_is_blank(row):
            return row_index
    return min_row


def _validate_excel_headers(headers: list[str]) -> None:
    if not headers:
        raise ValueError("Excel records 模式需要至少一个表头。")
    blank_indexes = [index + 1 for index, header in enumerate(headers) if not header]
    if blank_indexes:
        raise ValueError(f"Excel 表头不能为空，空表头列序号：{blank_indexes}")
    duplicates = sorted({header for header in headers if headers.count(header) > 1})
    if duplicates:
        raise ValueError(f"Excel 表头重复：{', '.join(duplicates)}")


def _excel_row_is_blank(row: Any) -> bool:
    for value in row:
        if value is None:
            continue
        if isinstance(value, str) and not value.strip():
            continue
        return False
    return True


def _ensure_max_rows(rows: list[Any], max_rows: int) -> None:
    if len(rows) > max_rows:
        raise ValueError(f"Excel 读取行数超过 max_rows={max_rows}，请缩小 range 或提高 max_rows。")


def _optional_positive_int(value: Any, *, field_name: str) -> int | None:
    if value in (None, ""):
        return None
    if not isinstance(value, int) or value < 1:
        raise ValueError(f"Excel {field_name} 必须是大于 0 的整数。")
    return value


def _optional_nonnegative_int(value: Any, *, field_name: str) -> int | None:
    if value in (None, ""):
        return None
    if not isinstance(value, int) or value < 0:
        raise ValueError(f"Excel {field_name} 必须是大于或等于 0 的整数。")
    return value


def _excel_row_window(
    start_row: int,
    max_row: int,
    *,
    offset_rows: int,
    limit_rows: int | None,
    preview_rows: int | None,
) -> tuple[int, int, bool]:
    effective_start = start_row + offset_rows
    effective_max = max_row
    limit_values = [value for value in (limit_rows, preview_rows) if value is not None]
    if limit_values:
        effective_max = min(effective_max, effective_start + min(limit_values) - 1)
    if effective_start > max_row:
        return effective_start, max_row, offset_rows > 0
    return effective_start, effective_max, offset_rows > 0 or effective_max < max_row


def _apply_excel_read_window_meta(
    meta: dict[str, Any],
    *,
    max_cells: int,
    offset_rows: int,
    limit_rows: int | None,
    preview_rows: int | None,
    truncated: bool,
) -> None:
    meta["truncated"] = bool(truncated)
    meta["max_cells"] = max_cells
    meta["offset_rows"] = offset_rows
    if limit_rows is not None:
        meta["limit_rows"] = limit_rows
    if preview_rows is not None:
        meta["preview_rows"] = preview_rows


def _ensure_max_cells(min_col: int, min_row: int, max_col: int, max_row: int, max_cells: int) -> None:
    if max_cells < 1:
        raise ValueError("Excel max_cells 必须大于 0。")
    if max_row < min_row or max_col < min_col:
        return
    _ensure_max_cell_count(max_col - min_col + 1, max_row - min_row + 1, max_cells)


def _ensure_max_cell_count(column_count: int, row_count: int, max_cells: int) -> None:
    if max_cells < 1:
        raise ValueError("Excel max_cells 必须大于 0。")
    if column_count <= 0 or row_count <= 0:
        return
    cell_count = column_count * row_count
    if cell_count > max_cells:
        raise ValueError(
            f"Excel 读取区域包含 {cell_count} 个单元格，超过 max_cells={max_cells}。"
            "请缩小 range、指定 sheet/preview_rows，或显式提高 max_cells。"
        )


def _excel_meta(path: Any, workbook: Any, worksheet: Any, range_text: str, headers: list[str], value: Any) -> dict[str, Any]:
    if isinstance(value, list):
        row_count = len(value)
    elif isinstance(value, dict):
        row_count = len(value)
    else:
        row_count = 0
    column_count = len(headers)
    if not column_count and isinstance(value, list) and value and isinstance(value[0], list):
        column_count = max(len(row) for row in value)
    return {
        "type": "excel",
        "path": str(path),
        "sheets": list(workbook.sheetnames),
        "sheet": worksheet.title,
        "range": range_text,
        "headers": headers,
        "row_count": row_count,
        "column_count": column_count,
    }


def _excel_range_text(min_col: int, min_row: int, max_col: int, max_row: int, get_column_letter: Any) -> str:
    if max_row < min_row or max_col < min_col:
        return ""
    return f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{max_row}"


def _json_safe_excel_value(value: Any, *, date_format: str = "iso") -> Any:
    if isinstance(value, (datetime, date, time)):
        if date_format == "text":
            return str(value)
        return value.isoformat()
    if isinstance(value, timedelta):
        return value.total_seconds()
    if isinstance(value, Decimal):
        return float(value)
    if isinstance(value, (str, int, float, bool)) or value is None:
        return value
    return str(value)


def _prepare_excel_worksheet(workbook: Any, raw_sheet: Any, *, write_mode: str, template: bool) -> Any:
    if isinstance(raw_sheet, int) and workbook.worksheets:
        sheet_name = workbook.worksheets[raw_sheet].title
    elif raw_sheet not in (None, ""):
        sheet_name = str(raw_sheet)
    elif template:
        sheet_name = workbook.worksheets[0].title
    else:
        sheet_name = "Sheet1"

    if write_mode in {"create", "replace_sheet"}:
        if sheet_name in workbook.sheetnames:
            old_sheet = workbook[sheet_name]
            index = workbook.worksheets.index(old_sheet)
            new_sheet = workbook.create_sheet(title=sheet_name, index=index)
            workbook.remove(old_sheet)
            return new_sheet
        if not template and len(workbook.worksheets) == 1 and _worksheet_is_empty(workbook.worksheets[0]):
            worksheet = workbook.worksheets[0]
            worksheet.title = sheet_name
            return worksheet
        return workbook.create_sheet(title=sheet_name)

    if write_mode in {"append_rows", "overlay_cells"}:
        if sheet_name in workbook.sheetnames:
            return workbook[sheet_name]
        return workbook.create_sheet(title=sheet_name)

    raise ValueError(f"不支持的 Excel write_mode：{write_mode}")


def _excel_write_sheet_steps(step: dict[str, Any]) -> list[dict[str, Any]]:
    raw_sheets = step.get("sheets")
    if raw_sheets is None:
        return [_normalize_excel_sheet_step(step)]
    if not isinstance(raw_sheets, list) or not raw_sheets:
        raise ValueError("write.type=excel.sheets 必须是非空数组。")
    sheet_steps: list[dict[str, Any]] = []
    for index, raw_sheet_step in enumerate(raw_sheets):
        if not isinstance(raw_sheet_step, dict):
            raise ValueError(f"write.type=excel.sheets[{index}] 必须是对象。")
        merged = {
            key: value
            for key, value in step.items()
            if key
            not in {
                "path",
                "template_path",
                "sheets",
                "sheet",
                "value",
                "rows",
                "cells",
            }
        }
        merged.update(raw_sheet_step)
        sheet_steps.append(_normalize_excel_sheet_step(merged))
    return sheet_steps


def _normalize_excel_sheet_step(step: dict[str, Any]) -> dict[str, Any]:
    normalized = dict(step)
    if "value" not in normalized and "rows" in normalized:
        normalized["value"] = normalized["rows"]
    if "value" not in normalized and "cells" not in normalized:
        raise ValueError("write.type=excel 需要 value、cells 或 sheets[].value/cells 之一。")
    return normalized


def _write_excel_sheet(
    workbook: Any,
    step: dict[str, Any],
    *,
    template: bool,
    get_column_letter: Any,
    range_boundaries: Any,
) -> None:
    step = _resolve_excel_named_range(workbook, step, range_boundaries)
    default_write_mode = (
        "overlay_cells"
        if template and (step.get("range") or ("cells" in step and "value" not in step))
        else ("replace_sheet" if template else "create")
    )
    write_mode = str(step.get("write_mode") or default_write_mode)
    worksheet = _prepare_excel_worksheet(
        workbook,
        step.get("sheet"),
        write_mode=write_mode,
        template=template,
    )

    data_range: tuple[int, int, int, int] | None = None
    if "value" in step:
        rows = _coerce_excel_rows(step.get("value"))
        start_row, start_col, max_row, max_col = _excel_write_bounds(step, range_boundaries)
        include_header = bool(step.get("include_header", True))
        data_range = _write_excel_rows(
            worksheet,
            rows,
            headers=step.get("headers"),
            append=write_mode == "append_rows",
            include_header=include_header,
            start_row=start_row,
            start_col=start_col,
            max_row=max_row,
            max_col=max_col,
            formula_columns=step.get("formula_columns"),
            get_column_letter=get_column_letter,
        )
        if data_range is not None and bool(step.get("copy_row_style", template and step.get("range"))):
            _copy_excel_template_row_style(
                worksheet,
                data_range,
                include_header=include_header,
                style_source_row=step.get("style_source_row"),
            )
        if data_range is not None and bool(step.get("extend_conditional_formatting", template and step.get("range"))):
            _extend_excel_conditional_formatting(worksheet, data_range, get_column_letter, range_boundaries)

    cells = step.get("cells")
    if isinstance(cells, dict):
        for address, value in cells.items():
            if not isinstance(address, str) or not EXCEL_A1_RE.match(address):
                raise ValueError(f"Excel cells 的 key 必须是 A1 单元格地址：{address}")
            row_number, column_number = _excel_cell_position(address, field_name="cells")
            _ensure_excel_cell_writable(worksheet, row_number, column_number)
            worksheet[address] = _excel_cell_value(value)

    _apply_excel_sheet_options(worksheet, step, get_column_letter, data_range=data_range)


def _worksheet_is_empty(worksheet: Any) -> bool:
    return worksheet.max_row == 1 and worksheet.max_column == 1 and worksheet["A1"].value is None


def _worksheet_last_nonempty_row(worksheet: Any) -> int:
    for row_number in range(worksheet.max_row, 0, -1):
        for column_number in range(1, worksheet.max_column + 1):
            value = worksheet.cell(row=row_number, column=column_number).value
            if value is None:
                continue
            if isinstance(value, str) and not value.strip():
                continue
            return row_number
    return 0


def _coerce_excel_rows(value: Any) -> list[Any]:
    if value is None:
        return []
    if isinstance(value, list):
        return value
    if isinstance(value, tuple):
        return list(value)
    if isinstance(value, dict):
        return [value]
    return [[value]]


def _write_excel_rows(
    worksheet: Any,
    rows: list[Any],
    *,
    headers: list[str] | None,
    append: bool,
    include_header: bool,
    start_row: int,
    start_col: int,
    max_row: int | None,
    max_col: int | None,
    formula_columns: Any,
    get_column_letter: Any,
) -> tuple[int, int, int, int] | None:
    existing_empty = _worksheet_is_empty(worksheet)
    target_row = _worksheet_last_nonempty_row(worksheet) + 1 if append and not existing_empty else start_row
    formula_specs = _normalize_excel_formula_columns(formula_columns)
    inferred_headers = _excel_write_headers(rows, headers, formula_columns=list(formula_specs.keys()))
    write_header = include_header and bool(inferred_headers) and (not append or existing_empty)
    _ensure_excel_write_fits_range(
        worksheet,
        rows,
        inferred_headers,
        target_row=target_row,
        start_col=start_col,
        max_row=max_row,
        max_col=max_col,
        include_header=write_header,
    )

    current_row = target_row
    max_width = 0
    if write_header:
        _write_excel_row(worksheet, current_row, inferred_headers, start_col=start_col)
        max_width = max(max_width, len(inferred_headers))
        current_row += 1

    for row in rows:
        if isinstance(row, dict):
            values = [
                _excel_formula_value(formula_specs[header], row, current_row, inferred_headers, start_col, get_column_letter)
                if header in formula_specs
                else row.get(header, "")
                for header in inferred_headers
            ]
        elif isinstance(row, (list, tuple)):
            if formula_specs:
                raise ValueError("write.type=excel.formula_columns 只支持字典行数组。")
            values = list(row)
        else:
            if formula_specs:
                raise ValueError("write.type=excel.formula_columns 只支持字典行数组。")
            values = [row]
        _write_excel_row(worksheet, current_row, values, start_col=start_col)
        max_width = max(max_width, len(values))
        current_row += 1
    if current_row == target_row or max_width == 0:
        return None
    return start_col, target_row, start_col + max_width - 1, current_row - 1


def _excel_write_headers(rows: list[Any], headers: list[str] | None, *, formula_columns: list[str] | None = None) -> list[str]:
    if headers:
        result = [str(header) for header in headers]
        for column in formula_columns or []:
            if column not in result:
                result.append(column)
        _validate_excel_headers(result)
        return result
    result: list[str] = []
    for row in rows:
        if not isinstance(row, dict):
            continue
        for key in row:
            text = str(key)
            if text not in result:
                result.append(text)
    for column in formula_columns or []:
        if column not in result:
            result.append(column)
    return result


def _write_excel_row(worksheet: Any, row_number: int, values: list[Any], *, start_col: int) -> None:
    for column_number, value in enumerate(values, start=start_col):
        worksheet.cell(row=row_number, column=column_number).value = _excel_cell_value(value)


def _excel_cell_value(value: Any) -> Any:
    if isinstance(value, Decimal):
        return float(value)
    if isinstance(value, (dict, list, tuple)):
        return json.dumps(value, ensure_ascii=False)
    if isinstance(value, (datetime, date, time, timedelta, str, int, float, bool)) or value is None:
        return value
    return str(value)


def _excel_cell_position(value: Any, *, field_name: str) -> tuple[int, int]:
    if not isinstance(value, str) or not value or ":" in value or not EXCEL_A1_RE.match(value):
        raise ValueError(f"Excel {field_name} 必须是 A1 风格单元格地址，例如 B3。")
    match = re.fullmatch(r"([A-Za-z]{1,3})([1-9][0-9]*)", value)
    if match is None:
        raise ValueError(f"Excel {field_name} 必须是 A1 风格单元格地址，例如 B3。")
    column_letter, row_text = match.groups()
    return int(row_text), _column_number(column_letter)


def _excel_write_bounds(step: dict[str, Any], range_boundaries: Any) -> tuple[int, int, int | None, int | None]:
    raw_range = step.get("range")
    if raw_range:
        range_text = str(raw_range)
        if not EXCEL_A1_RE.match(range_text) or ":" not in range_text:
            raise ValueError(f"Excel write.range 必须是 A1 风格范围，例如 B3:H20：{range_text}")
        min_col, min_row, max_col, max_row = range_boundaries(range_text)
        return min_row, min_col, max_row, max_col
    start_row, start_col = _excel_cell_position(step.get("start_cell") or "A1", field_name="start_cell")
    return start_row, start_col, None, None


def _resolve_excel_named_range(workbook: Any, step: dict[str, Any], range_boundaries: Any) -> dict[str, Any]:
    raw_name = step.get("named_range")
    if raw_name in (None, ""):
        return step
    if not isinstance(raw_name, str):
        raise ValueError("write.type=excel.named_range 必须是非空字符串。")
    if step.get("range"):
        raise ValueError("write.type=excel.named_range 不能和 range 同时使用。")

    defined_names = getattr(workbook, "defined_names", {})
    try:
        defined_name = defined_names[raw_name]
    except (KeyError, TypeError):
        defined_name = defined_names.get(raw_name) if hasattr(defined_names, "get") else None
    if defined_name is None:
        raise ValueError(f"Excel 模板中未找到命名区域：{raw_name}")

    destinations = list(getattr(defined_name, "destinations", []))
    if len(destinations) != 1:
        raise ValueError(f"Excel named_range 必须指向唯一单元格区域：{raw_name}")
    sheet_name, raw_range = destinations[0]
    range_text = str(raw_range).replace("$", "")
    if not EXCEL_A1_RE.match(range_text) or ":" not in range_text:
        raise ValueError(f"Excel named_range 必须指向 A1 范围：{raw_name} -> {raw_range}")
    range_boundaries(range_text)

    explicit_sheet = step.get("sheet")
    if explicit_sheet not in (None, "") and str(explicit_sheet) != str(sheet_name):
        raise ValueError(f"Excel named_range={raw_name} 位于 sheet={sheet_name}，不能同时指定 sheet={explicit_sheet}。")

    resolved = dict(step)
    resolved["sheet"] = str(sheet_name)
    resolved["range"] = range_text
    return resolved


def _ensure_excel_cell_writable(worksheet: Any, row_number: int, column_number: int) -> None:
    for merged_range in worksheet.merged_cells.ranges:
        if (
            merged_range.min_row <= row_number <= merged_range.max_row
            and merged_range.min_col <= column_number <= merged_range.max_col
            and (row_number != merged_range.min_row or column_number != merged_range.min_col)
        ):
            raise ValueError(
                f"Excel 写入目标 {worksheet.cell(row=row_number, column=column_number).coordinate} "
                f"位于合并单元格 {merged_range.coord} 内，且不是左上角单元格。"
            )


def _ensure_excel_range_writable(
    worksheet: Any,
    min_row: int,
    min_col: int,
    max_row: int,
    max_col: int,
) -> None:
    for row_number in range(min_row, max_row + 1):
        for column_number in range(min_col, max_col + 1):
            _ensure_excel_cell_writable(worksheet, row_number, column_number)


def _copy_excel_template_row_style(
    worksheet: Any,
    data_range: tuple[int, int, int, int],
    *,
    include_header: bool,
    style_source_row: Any,
) -> None:
    min_col, min_row, max_col, max_row = data_range
    first_data_row = min_row + 1 if include_header else min_row
    if first_data_row > max_row:
        return
    if style_source_row not in (None, ""):
        if not isinstance(style_source_row, int) or style_source_row < 1:
            raise ValueError("write.type=excel.style_source_row 必须是大于 0 的整数。")
        source_row = style_source_row
    else:
        source_row = first_data_row
    for target_row in range(first_data_row, max_row + 1):
        if target_row == source_row:
            continue
        for column_number in range(min_col, max_col + 1):
            _copy_excel_cell_style(
                worksheet.cell(row=source_row, column=column_number),
                worksheet.cell(row=target_row, column=column_number),
            )


def _copy_excel_cell_style(source_cell: Any, target_cell: Any) -> None:
    if not getattr(source_cell, "has_style", False):
        return
    target_cell.font = copy(source_cell.font)
    target_cell.fill = copy(source_cell.fill)
    target_cell.border = copy(source_cell.border)
    target_cell.alignment = copy(source_cell.alignment)
    target_cell.number_format = source_cell.number_format
    target_cell.protection = copy(source_cell.protection)


def _extend_excel_conditional_formatting(
    worksheet: Any,
    data_range: tuple[int, int, int, int],
    get_column_letter: Any,
    range_boundaries: Any,
) -> None:
    data_min_col, data_min_row, data_max_col, data_max_row = data_range
    for conditional_formatting in list(worksheet.conditional_formatting):
        for raw_range in str(conditional_formatting.sqref).split():
            source_ref = raw_range.replace("$", "")
            if not EXCEL_A1_RE.match(source_ref):
                continue
            min_col, min_row, max_col, max_row = range_boundaries(source_ref)
            source_range = (min_col, min_row, max_col, max_row)
            if not _excel_ranges_intersect(source_range, data_range):
                continue
            target_ref = _excel_range_text(
                max(min_col, data_min_col),
                max(min_row, data_min_row),
                min(max_col, data_max_col),
                data_max_row,
                get_column_letter,
            )
            if not target_ref or source_ref.upper() == target_ref.upper():
                continue
            for rule in getattr(conditional_formatting, "rules", []):
                worksheet.conditional_formatting.add(target_ref, copy(rule))


def _excel_ranges_intersect(left: tuple[int, int, int, int], right: tuple[int, int, int, int]) -> bool:
    left_min_col, left_min_row, left_max_col, left_max_row = left
    right_min_col, right_min_row, right_max_col, right_max_row = right
    return not (
        left_max_col < right_min_col
        or right_max_col < left_min_col
        or left_max_row < right_min_row
        or right_max_row < left_min_row
    )


def _ensure_excel_write_fits_range(
    worksheet: Any,
    rows: list[Any],
    headers: list[str],
    *,
    target_row: int,
    start_col: int,
    max_row: int | None,
    max_col: int | None,
    include_header: bool,
) -> None:
    if max_row is None and max_col is None:
        return
    height = len(rows) + (1 if include_header else 0)
    widths = [len(headers)] if headers else []
    for row in rows:
        if isinstance(row, dict):
            widths.append(len(headers))
        elif isinstance(row, (list, tuple)):
            widths.append(len(row))
        else:
            widths.append(1)
    width = max(widths or [0])
    if max_row is not None and target_row + height - 1 > max_row:
        raise ValueError("write.type=excel.range 行数不足，数据会超出指定区域。")
    if max_col is not None and start_col + width - 1 > max_col:
        raise ValueError("write.type=excel.range 列数不足，数据会超出指定区域。")
    if height > 0 and width > 0:
        _ensure_excel_range_writable(
            worksheet,
            target_row,
            start_col,
            target_row + height - 1,
            start_col + width - 1,
        )


def _normalize_excel_formula_columns(value: Any) -> dict[str, str]:
    if value in (None, ""):
        return {}
    if not isinstance(value, dict):
        raise ValueError("write.type=excel.formula_columns 必须是对象。")
    result: dict[str, str] = {}
    for column, raw_spec in value.items():
        if not isinstance(column, str) or not column:
            raise ValueError("write.type=excel.formula_columns 的列名必须是非空字符串。")
        if isinstance(raw_spec, str) and raw_spec:
            result[column] = raw_spec
            continue
        if isinstance(raw_spec, dict) and isinstance(raw_spec.get("formula"), str) and raw_spec["formula"]:
            result[column] = str(raw_spec["formula"])
            continue
        raise ValueError("write.type=excel.formula_columns 每项必须是公式字符串或包含 formula 的对象。")
    return result


def _excel_formula_value(
    formula: str,
    row: dict[str, Any],
    row_number: int,
    headers: list[str],
    start_col: int,
    get_column_letter: Any,
) -> str:
    header_cells = {
        header: f"{get_column_letter(start_col + index)}{row_number}"
        for index, header in enumerate(headers)
        if header in row
    }
    value = formula.replace("{row}", str(row_number))
    for header, address in header_cells.items():
        value = value.replace("{" + header + "}", address)
    return value if value.startswith("=") else f"={value}"


def _excel_effective_data_range(
    worksheet: Any,
    data_range: tuple[int, int, int, int] | None,
) -> tuple[int, int, int, int]:
    if data_range is not None:
        return data_range
    return 1, 1, worksheet.max_column, worksheet.max_row


def _apply_excel_sheet_options(
    worksheet: Any,
    step: dict[str, Any],
    get_column_letter: Any,
    *,
    data_range: tuple[int, int, int, int] | None,
) -> None:
    range_min_col, range_min_row, range_max_col, range_max_row = _excel_effective_data_range(
        worksheet,
        data_range,
    )
    if bool(step.get("freeze_header", False)):
        worksheet.freeze_panes = f"{get_column_letter(range_min_col)}{range_min_row + 1}"
    if bool(step.get("auto_filter", False)) and range_max_row >= range_min_row and range_max_col >= range_min_col:
        worksheet.auto_filter.ref = _excel_range_text(range_min_col, range_min_row, range_max_col, range_max_row, get_column_letter)
    _apply_excel_column_widths(worksheet, step.get("column_widths"), get_column_letter, header_row=range_min_row)
    _apply_excel_number_formats(worksheet, step.get("number_format"), get_column_letter, header_row=range_min_row)
    if bool(step.get("table", False)) and range_max_row >= range_min_row + 1 and range_max_col >= range_min_col:
        _add_excel_table(worksheet, step, get_column_letter, data_range=(range_min_col, range_min_row, range_max_col, range_max_row))


def _apply_excel_column_widths(worksheet: Any, column_widths: Any, get_column_letter: Any, *, header_row: int) -> None:
    if not isinstance(column_widths, dict):
        return
    header_map = _excel_header_column_map(worksheet, header_row=header_row)
    for key, width in column_widths.items():
        column_letter = _excel_column_letter(str(key), header_map, get_column_letter)
        worksheet.column_dimensions[column_letter].width = float(width)


def _apply_excel_number_formats(worksheet: Any, number_format: Any, get_column_letter: Any, *, header_row: int) -> None:
    if not isinstance(number_format, dict):
        return
    header_map = _excel_header_column_map(worksheet, header_row=header_row)
    for key, fmt in number_format.items():
        column_letter = _excel_column_letter(str(key), header_map, get_column_letter)
        column_index = header_map.get(str(key), None)
        if column_index is None and column_letter.isalpha():
            column_index = _column_number(column_letter)
        if column_index is None:
            continue
        for row_number in range(header_row + 1, worksheet.max_row + 1):
            worksheet.cell(row=row_number, column=column_index).number_format = str(fmt)


def _excel_header_column_map(worksheet: Any, *, header_row: int = 1) -> dict[str, int]:
    mapping: dict[str, int] = {}
    if worksheet.max_row < header_row:
        return mapping
    for column_number in range(1, worksheet.max_column + 1):
        value = worksheet.cell(row=header_row, column=column_number).value
        if value not in (None, ""):
            mapping[str(value)] = column_number
    return mapping


def _excel_column_letter(key: str, header_map: dict[str, int], get_column_letter: Any) -> str:
    if key in header_map:
        return get_column_letter(header_map[key])
    if re.fullmatch(r"[A-Za-z]{1,3}", key):
        return key.upper()
    raise ValueError(f"无法定位 Excel 列：{key}")


def _column_number(column_letter: str) -> int:
    number = 0
    for char in column_letter.upper():
        number = number * 26 + (ord(char) - ord("A") + 1)
    return number


def _add_excel_table(
    worksheet: Any,
    step: dict[str, Any],
    get_column_letter: Any,
    *,
    data_range: tuple[int, int, int, int] | None,
) -> None:
    excel = _load_openpyxl()
    min_col, min_row, max_col, max_row = _excel_effective_data_range(worksheet, data_range)
    ref = _excel_range_text(min_col, min_row, max_col, max_row, get_column_letter)
    table_name = _safe_excel_table_name(str(step.get("table_name") or worksheet.title or "Table1"))
    existing = {
        table.name
        for workbook_sheet in worksheet.parent.worksheets
        for table in workbook_sheet.tables.values()
    }
    base_name = table_name
    index = 1
    while table_name in existing:
        index += 1
        table_name = f"{base_name}{index}"
    table = excel["Table"](displayName=table_name, ref=ref)
    table.tableStyleInfo = excel["TableStyleInfo"](
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    worksheet.add_table(table)


def _safe_excel_table_name(value: str) -> str:
    name = re.sub(r"[^A-Za-z0-9_]", "_", value.strip()) or "Table1"
    if name[0].isdigit():
        name = f"Table_{name}"
    return name[:255]
