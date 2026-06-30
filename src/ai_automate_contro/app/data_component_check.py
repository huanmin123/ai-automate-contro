from __future__ import annotations

import json
import platform
import tempfile
import time
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Callable

from ai_automate_contro.engine.executor import execute_plan
from ai_automate_contro.plans.loader import load_plan
from ai_automate_contro.plans.validator import validate_plan_file


@dataclass(frozen=True)
class DataRegressionCase:
    name: str
    plan_path: str
    evidence_check: Callable[[Path, float], list[dict[str, Any]]]


DATA_REGRESSION_CASES = (
    DataRegressionCase(
        "write-excel",
        "test-plans/basic/write-excel/plan.json",
        lambda root, started_at: _write_excel_evidence(root, started_at),
    ),
    DataRegressionCase(
        "load-excel",
        "test-plans/data-driven/load-excel/plan.json",
        lambda root, started_at: _load_excel_evidence(root, started_at),
    ),
    DataRegressionCase(
        "table-transform",
        "test-plans/data-driven/table-transform/plan.json",
        lambda root, started_at: _table_transform_evidence(root, started_at),
    ),
    DataRegressionCase(
        "table-cleaning",
        "test-plans/data-driven/table-cleaning/plan.json",
        lambda root, started_at: _table_cleaning_evidence(root, started_at),
    ),
    DataRegressionCase(
        "enterprise-cookbook",
        "test-plans/data-driven/enterprise-cookbook/plan.json",
        lambda root, started_at: _enterprise_cookbook_evidence(root, started_at),
    ),
)

NEGATIVE_VALIDATION_CASES = (
    {
        "name": "excel-read-invalid-range",
        "expected_message": "range 必须是 A1 风格单元格或范围",
        "plan": {
            "name": "excel-read-invalid-range",
            "automation_type": "browser",
            "steps": [
                {
                    "action": "read",
                    "type": "excel",
                    "path": "resources/source.xlsx",
                    "range": "not-a-range",
                    "save_as": "rows",
                }
            ],
        },
    },
    {
        "name": "excel-write-missing-data",
        "expected_message": "write.type=excel 需要 value、cells 或 sheets 之一",
        "plan": {
            "name": "excel-write-missing-data",
            "automation_type": "browser",
            "steps": [{"action": "write", "type": "excel", "path": "bad.xlsx"}],
        },
    },
    {
        "name": "excel-write-invalid-start-cell",
        "expected_message": "start_cell 必须是 A1 风格单元格地址",
        "plan": {
            "name": "excel-write-invalid-start-cell",
            "automation_type": "browser",
            "steps": [
                {
                    "action": "write",
                    "type": "excel",
                    "path": "bad.xlsx",
                    "start_cell": "A1:B2",
                    "value": [{"a": 1}],
                }
            ],
        },
    },
    {
        "name": "excel-read-empty-sheets",
        "expected_message": "sheets 必须是非空数组",
        "plan": {
            "name": "excel-read-empty-sheets",
            "automation_type": "browser",
            "steps": [
                {
                    "action": "read",
                    "type": "excel",
                    "path": "resources/source.xlsx",
                    "sheets": [],
                    "save_as": "workbook",
                }
            ],
        },
    },
    {
        "name": "excel-read-invalid-max-cells",
        "expected_message": "max_cells 必须大于或等于 1",
        "plan": {
            "name": "excel-read-invalid-max-cells",
            "automation_type": "browser",
            "steps": [
                {
                    "action": "read",
                    "type": "excel",
                    "path": "resources/source.xlsx",
                    "max_cells": 0,
                    "save_as": "rows",
                }
            ],
        },
    },
    {
        "name": "excel-read-invalid-offset-rows",
        "expected_message": "offset_rows 必须大于或等于 0",
        "plan": {
            "name": "excel-read-invalid-offset-rows",
            "automation_type": "browser",
            "steps": [
                {
                    "action": "read",
                    "type": "excel",
                    "path": "resources/source.xlsx",
                    "offset_rows": -1,
                    "save_as": "rows",
                }
            ],
        },
    },
    {
        "name": "excel-write-invalid-formula-column",
        "expected_message": "formula_columns.合计 必须是公式字符串或包含 formula 的对象",
        "plan": {
            "name": "excel-write-invalid-formula-column",
            "automation_type": "browser",
            "steps": [
                {
                    "action": "write",
                    "type": "excel",
                    "path": "bad.xlsx",
                    "value": [{"金额": 10}],
                    "formula_columns": {"合计": {"expression": "{金额}"}},
                }
            ],
        },
    },
    {
        "name": "excel-write-invalid-range-cell",
        "expected_message": "range 必须是 A1 风格范围",
        "plan": {
            "name": "excel-write-invalid-range-cell",
            "automation_type": "browser",
            "steps": [
                {
                    "action": "write",
                    "type": "excel",
                    "path": "bad.xlsx",
                    "range": "B4",
                    "value": [{"金额": 10}],
                }
            ],
        },
    },
    {
        "name": "excel-write-named-range-conflict",
        "expected_message": "write.type=excel.named_range 不能和 range 同时使用",
        "plan": {
            "name": "excel-write-named-range-conflict",
            "automation_type": "browser",
            "steps": [
                {
                    "action": "write",
                    "type": "excel",
                    "path": "bad.xlsx",
                    "named_range": "DetailArea",
                    "range": "B4:F20",
                    "value": [{"金额": 10}],
                }
            ],
        },
    },
    {
        "name": "table-sort-descending-length",
        "expected_message": "table.sort.descending 数组长度必须与 by 一致",
        "plan": {
            "name": "table-sort-descending-length",
            "automation_type": "browser",
            "variables": {"rows": [{"a": 1, "b": 2}]},
            "steps": [
                {
                    "action": "table",
                    "type": "sort",
                    "source": "{{rows}}",
                    "by": ["a", "b"],
                    "descending": [True],
                    "save_as": "sorted_rows",
                }
            ],
        },
    },
    {
        "name": "table-join-missing-key",
        "expected_message": "table.join 需要 on，或同时提供 left_on 和 right_on",
        "plan": {
            "name": "table-join-missing-key",
            "automation_type": "browser",
            "variables": {"left": [], "right": []},
            "steps": [
                {
                    "action": "table",
                    "type": "join",
                    "source": "{{left}}",
                    "right": "{{right}}",
                    "save_as": "joined",
                }
            ],
        },
    },
    {
        "name": "table-pivot-invalid-agg",
        "expected_message": "agg 不支持的取值",
        "plan": {
            "name": "table-pivot-invalid-agg",
            "automation_type": "browser",
            "variables": {"rows": [{"部门": "财务", "级别": "P2", "工资": 12000}]},
            "steps": [
                {
                    "action": "table",
                    "type": "pivot",
                    "source": "{{rows}}",
                    "index": "部门",
                    "columns": "级别",
                    "values": "工资",
                    "agg": "median",
                    "save_as": "pivot_rows",
                }
            ],
        },
    },
    {
        "name": "table-fuzzy-lookup-invalid-threshold",
        "expected_message": "table.fuzzy_lookup.threshold 必须在 0 到 1 之间",
        "plan": {
            "name": "table-fuzzy-lookup-invalid-threshold",
            "automation_type": "browser",
            "variables": {"left": [], "right": []},
            "steps": [
                {
                    "action": "table",
                    "type": "fuzzy_lookup",
                    "source": "{{left}}",
                    "right": "{{right}}",
                    "on": "姓名",
                    "threshold": 1.5,
                    "save_as": "looked_up",
                }
            ],
        },
    },
    {
        "name": "table-lookup-missing-key",
        "expected_message": "table.lookup 需要 on，或同时提供 left_on 和 right_on",
        "plan": {
            "name": "table-lookup-missing-key",
            "automation_type": "browser",
            "variables": {"left": [], "right": []},
            "steps": [
                {
                    "action": "table",
                    "type": "lookup",
                    "source": "{{left}}",
                    "right": "{{right}}",
                    "save_as": "looked_up",
                }
            ],
        },
    },
)


def self_check_data_components(project_root: str | Path) -> dict[str, Any]:
    root = Path(project_root).resolve()
    positive_cases = [_run_regression_case(root, regression_case) for regression_case in DATA_REGRESSION_CASES]
    dynamic_cases = [
        _run_template_style_case(root),
        _run_large_excel_paging_case(root),
        _run_template_runtime_negative_cases(root),
    ]
    negative_cases = _run_negative_validation_cases(root)
    positive_ok = all(case["ok"] for case in positive_cases)
    dynamic_ok = all(case["ok"] for case in dynamic_cases)
    negative_ok = all(case["ok"] for case in negative_cases)
    return {
        "ok": positive_ok and dynamic_ok and negative_ok,
        "check": "data_components",
        "project_root": str(root),
        "checks": [
            {
                "name": "data_plan_regression_matrix",
                "ok": positive_ok,
                "cases": positive_cases,
            },
            {
                "name": "excel_template_style_regression",
                "ok": dynamic_ok,
                "cases": dynamic_cases,
            },
            {
                "name": "data_negative_validation",
                "ok": negative_ok,
                "cases": negative_cases,
            },
        ],
        "commands": {"run": f"python {_cplan_script_path()} self-check data-components"},
    }


def _run_regression_case(project_root: Path, regression_case: DataRegressionCase) -> dict[str, Any]:
    plan_path = project_root / regression_case.plan_path
    validation = validate_plan_file(plan_path, project_root)
    if not validation.ok:
        return _validation_failed_case(regression_case.name, plan_path, validation)

    try:
        plan = load_plan(plan_path)
        started_at = time.time()
        result = execute_plan(
            plan,
            project_root,
            plan_path=plan_path,
            run_name=f"data-components-{regression_case.name}",
            run_context_handler=_disable_run_log_echo,
        )
    except Exception as error:
        return _run_failed_case(regression_case.name, plan_path, error)

    try:
        evidence = regression_case.evidence_check(project_root, started_at)
    except Exception as error:
        evidence = [{"name": "evidence_read", "ok": False, "error": str(error), "error_type": type(error).__name__}]

    return {
        "name": regression_case.name,
        "ok": result.status == "passed" and all(item["ok"] for item in evidence),
        "plan_path": str(plan_path),
        "validation_ok": True,
        "run_ok": result.status == "passed",
        "output_dir": result.output_dir,
        "evidence": evidence,
    }


def _run_template_style_case(project_root: Path) -> dict[str, Any]:
    with tempfile.TemporaryDirectory(prefix="data-components-excel-template-") as raw_temp_dir:
        package_dir = Path(raw_temp_dir) / "excel-template"
        resources_dir = package_dir / "resources"
        resources_dir.mkdir(parents=True, exist_ok=True)
        template_path = resources_dir / "template.xlsx"
        _create_template_workbook(template_path)
        plan_path = package_dir / "plan.json"
        plan_path.write_text(json.dumps(_template_style_plan(), ensure_ascii=False, indent=2) + "\n", encoding="utf-8")

        validation = validate_plan_file(plan_path, project_root)
        if not validation.ok:
            return _validation_failed_case("excel-template-style", plan_path, validation)
        try:
            plan = load_plan(plan_path)
            started_at = time.time()
            result = execute_plan(
                plan,
                project_root,
                plan_path=plan_path,
                run_name="data-components-excel-template-style",
                run_context_handler=_disable_run_log_echo,
            )
        except Exception as error:
            return _run_failed_case("excel-template-style", plan_path, error)

        output_path = package_dir / "output" / "excel" / "report.xlsx"
        preview_path = package_dir / "output" / "json" / "template-preview.json"
        evidence = [
            _expect("run_passed", result.status == "passed"),
            _expect("report_fresh", _file_nonempty_after(output_path, started_at)),
            _expect("preview_json_fresh", _file_nonempty_after(preview_path, started_at)),
        ]
        try:
            evidence.extend(_template_style_evidence(output_path))
        except Exception as error:
            evidence.append({"name": "template_evidence_read", "ok": False, "error": str(error), "error_type": type(error).__name__})
        try:
            evidence.extend(_template_preview_evidence(preview_path))
        except Exception as error:
            evidence.append({"name": "template_preview_read", "ok": False, "error": str(error), "error_type": type(error).__name__})
        return {
            "name": "excel-template-style",
            "ok": result.status == "passed" and all(item["ok"] for item in evidence),
            "plan_path": str(plan_path),
            "output_dir": result.output_dir,
            "evidence": evidence,
        }


def _run_large_excel_paging_case(project_root: Path) -> dict[str, Any]:
    with tempfile.TemporaryDirectory(prefix="data-components-large-excel-") as raw_temp_dir:
        package_dir = Path(raw_temp_dir) / "large-excel-paging"
        resources_dir = package_dir / "resources"
        resources_dir.mkdir(parents=True, exist_ok=True)
        workbook_path = resources_dir / "large.xlsx"
        _create_large_excel_workbook(workbook_path, row_count=12000)
        plan_path = package_dir / "plan.json"
        plan_path.write_text(json.dumps(_large_excel_paging_plan(), ensure_ascii=False, indent=2) + "\n", encoding="utf-8")

        validation = validate_plan_file(plan_path, project_root)
        if not validation.ok:
            return _validation_failed_case("large-excel-paging", plan_path, validation)
        try:
            plan = load_plan(plan_path)
            started_at = time.time()
            result = execute_plan(
                plan,
                project_root,
                plan_path=plan_path,
                run_name="data-components-large-excel-paging",
                run_context_handler=_disable_run_log_echo,
            )
        except Exception as error:
            return _run_failed_case("large-excel-paging", plan_path, error)

        output_path = package_dir / "output" / "json" / "large-excel-paging.json"
        evidence = [
            _expect("run_passed", result.status == "passed"),
            _expect("json_fresh", _file_nonempty_after(output_path, started_at)),
        ]
        try:
            evidence.extend(_large_excel_paging_evidence(output_path))
        except Exception as error:
            evidence.append({"name": "large_excel_paging_evidence_read", "ok": False, "error": str(error), "error_type": type(error).__name__})
        return {
            "name": "large-excel-paging",
            "ok": result.status == "passed" and all(item["ok"] for item in evidence),
            "plan_path": str(plan_path),
            "output_dir": result.output_dir,
            "evidence": evidence,
        }


def _run_template_runtime_negative_cases(project_root: Path) -> dict[str, Any]:
    cases: list[dict[str, Any]] = []
    with tempfile.TemporaryDirectory(prefix="data-components-template-negative-") as raw_temp_dir:
        temp_dir = Path(raw_temp_dir)
        for raw_case in _template_runtime_negative_case_specs():
            cases.append(_run_template_runtime_negative_case(project_root, temp_dir, raw_case))
    return {
        "name": "excel-template-runtime-negative",
        "ok": all(case["ok"] for case in cases),
        "cases": cases,
    }


def _run_template_runtime_negative_case(project_root: Path, temp_dir: Path, raw_case: dict[str, Any]) -> dict[str, Any]:
    name = str(raw_case["name"])
    expected_message = str(raw_case["expected_message"])
    package_dir = temp_dir / name
    resources_dir = package_dir / "resources"
    resources_dir.mkdir(parents=True, exist_ok=True)
    _create_template_workbook(resources_dir / "template.xlsx")
    plan_path = package_dir / "plan.json"
    plan_path.write_text(json.dumps(raw_case["plan"], ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    validation = validate_plan_file(plan_path, project_root)
    if not validation.ok:
        return _validation_failed_case(name, plan_path, validation)
    try:
        plan = load_plan(plan_path)
        result = execute_plan(
            plan,
            project_root,
            plan_path=plan_path,
            run_name=f"data-components-{name}",
            run_context_handler=_disable_run_log_echo,
        )
    except Exception as error:
        error_text = str(error)
        return {
            "name": name,
            "ok": expected_message in error_text,
            "plan_path": str(plan_path),
            "validation_ok": True,
            "run_ok": False,
            "matched": expected_message in error_text,
            "error": error_text,
            "error_type": type(error).__name__,
        }
    return {
        "name": name,
        "ok": False,
        "plan_path": str(plan_path),
        "validation_ok": True,
        "run_ok": result.status == "passed",
        "expected_message": expected_message,
        "error": "",
    }


def _run_negative_validation_cases(project_root: Path) -> list[dict[str, Any]]:
    results: list[dict[str, Any]] = []
    with tempfile.TemporaryDirectory(prefix="data-components-negative-") as raw_temp_dir:
        temp_dir = Path(raw_temp_dir)
        for raw_case in NEGATIVE_VALIDATION_CASES:
            results.append(_run_negative_validation_case(project_root, temp_dir, raw_case))
    return results


def _run_negative_validation_case(project_root: Path, temp_dir: Path, raw_case: dict[str, Any]) -> dict[str, Any]:
    name = str(raw_case["name"])
    expected_message = str(raw_case["expected_message"])
    package_dir = temp_dir / name
    package_dir.mkdir(parents=True, exist_ok=True)
    plan_path = package_dir / "plan.json"
    plan_path.write_text(json.dumps(raw_case["plan"], ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    validation = validate_plan_file(plan_path, project_root)
    errors = [error.format() for error in validation.errors]
    matched = any(expected_message in error for error in errors)
    return {
        "name": name,
        "ok": (not validation.ok) and matched,
        "expected_message": expected_message,
        "validation_ok": validation.ok,
        "matched": matched,
        "errors": errors,
    }


def _template_style_plan() -> dict[str, Any]:
    return {
        "name": "excel-template-style",
        "automation_type": "browser",
        "steps": [
            {
                "action": "write",
                "type": "excel",
                "path": "report.xlsx",
                "template_path": "resources/template.xlsx",
                "named_range": "DetailArea",
                "value": [
                    {"姓名": "张三", "部门": "财务", "基本工资": 12000, "奖金": 1000},
                    {"姓名": "李四", "部门": "人事", "基本工资": 9000, "奖金": 500},
                ],
                "formula_columns": {"实发": "={基本工资}+{奖金}"},
                "cells": {"B2": "2026-06", "F2": "已生成"},
                "copy_row_style": True,
                "style_source_row": 5,
                "extend_conditional_formatting": True,
            },
            {
                "action": "read",
                "type": "excel",
                "path": "output/excel/report.xlsx",
                "sheet": "模板",
                "range": "B4:F20",
                "offset_rows": 1,
                "limit_rows": 1,
                "preview_rows": 1,
                "max_cells": 20,
                "save_as": "preview_rows",
                "save_meta_as": "preview_meta",
            },
            {
                "action": "write",
                "type": "json",
                "path": "template-preview.json",
                "value": {"rows": "{{preview_rows}}", "meta": "{{preview_meta}}"},
            },
        ],
    }


def _large_excel_paging_plan() -> dict[str, Any]:
    return {
        "name": "large-excel-paging",
        "automation_type": "browser",
        "steps": [
            {
                "action": "read",
                "type": "excel",
                "path": "resources/large.xlsx",
                "sheet": "明细",
                "range": "A1:H12001",
                "offset_rows": 1000,
                "limit_rows": 25,
                "max_cells": 208,
                "save_as": "page_rows",
                "save_meta_as": "page_meta",
            },
            {
                "action": "read",
                "type": "excel",
                "path": "resources/large.xlsx",
                "sheet": "明细",
                "range": "A1:H12001",
                "offset_rows": 11990,
                "limit_rows": 25,
                "max_cells": 208,
                "save_as": "tail_rows",
                "save_meta_as": "tail_meta",
            },
            {
                "action": "write",
                "type": "json",
                "path": "large-excel-paging.json",
                "value": {
                    "page": "{{page_rows}}",
                    "page_meta": "{{page_meta}}",
                    "tail": "{{tail_rows}}",
                    "tail_meta": "{{tail_meta}}",
                },
            },
        ],
    }


def _create_template_workbook(path: Path) -> None:
    from openpyxl import Workbook
    from openpyxl.formatting.rule import CellIsRule
    from openpyxl.styles import Font, PatternFill
    from openpyxl.workbook.defined_name import DefinedName

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "模板"
    worksheet["A1"] = "人员模板"
    worksheet.merge_cells("A1:F1")
    worksheet["A1"].fill = PatternFill(fill_type="solid", fgColor="FF4472C4")
    worksheet["A1"].font = Font(color="FFFFFFFF", bold=True)
    worksheet["B4"] = "姓名"
    worksheet["C4"] = "部门"
    worksheet["D4"] = "基本工资"
    worksheet["E4"] = "奖金"
    worksheet["F4"] = "实发"
    header_fill = PatternFill(fill_type="solid", fgColor="FFD9EAF7")
    data_fill = PatternFill(fill_type="solid", fgColor="FFFFF2CC")
    high_value_fill = PatternFill(fill_type="solid", fgColor="FFFFC7CE")
    for cell in worksheet["B4:F4"][0]:
        cell.fill = header_fill
        cell.font = Font(bold=True)
    for cell in worksheet["B5:F5"][0]:
        cell.fill = data_fill
    worksheet["A3"] = "保留行"
    worksheet["F5"] = "=D5+E5"
    worksheet.conditional_formatting.add(
        "F5:F5",
        CellIsRule(operator="greaterThan", formula=["10000"], fill=high_value_fill),
    )
    defined_name = DefinedName("DetailArea", attr_text=f"'{worksheet.title}'!$B$4:$F$7")
    if hasattr(workbook.defined_names, "add"):
        workbook.defined_names.add(defined_name)
    else:
        workbook.defined_names.append(defined_name)
    workbook.save(path)


def _create_large_excel_workbook(path: Path, *, row_count: int) -> None:
    from openpyxl import Workbook

    workbook = Workbook(write_only=True)
    worksheet = workbook.create_sheet("明细")
    worksheet.append(["序号", "单号", "部门", "金额", "税额", "状态", "日期", "备注"])
    for index in range(1, row_count + 1):
        department = "财务" if index % 2 else "人事"
        worksheet.append(
            [
                index,
                f"TXN-{index:05d}",
                department,
                index * 10,
                round(index * 0.6, 2),
                "有效",
                f"2026-06-{(index % 28) + 1:02d}",
                "",
            ]
        )
    workbook.save(path)


def _template_style_evidence(path: Path) -> list[dict[str, Any]]:
    from openpyxl import load_workbook

    workbook = load_workbook(path)
    try:
        worksheet = workbook["模板"]
        return [
            _expect("template_sheet_preserved", "模板" in workbook.sheetnames),
            _expect("title_cell_b2", worksheet["B2"].value == "2026-06"),
            _expect("summary_cell_f2", worksheet["F2"].value == "已生成"),
            _expect("range_header_b4", worksheet["B4"].value == "姓名"),
            _expect("range_header_f4", worksheet["F4"].value == "实发"),
            _expect("range_row_b5", worksheet["B5"].value == "张三"),
            _expect("range_row_b6", worksheet["B6"].value == "李四"),
            _expect("formula_column_f5", worksheet["F5"].value == "=D5+E5"),
            _expect("formula_column_f6", worksheet["F6"].value == "=D6+E6"),
            _expect("existing_value_preserved", worksheet["A3"].value == "保留行"),
            _expect("header_fill_preserved", worksheet["A1"].fill.fgColor.rgb == "FF4472C4"),
            _expect("named_range_preserved", "DetailArea" in workbook.defined_names),
            _expect("row_style_copied", worksheet["B6"].fill.fgColor.rgb == "FFFFF2CC"),
            _expect("conditional_formatting_extended", _worksheet_has_conditional_range(worksheet, "F5:F6")),
        ]
    finally:
        workbook.close()


def _template_preview_evidence(path: Path) -> list[dict[str, Any]]:
    payload = _read_json(path)
    rows = payload.get("rows", []) if isinstance(payload, dict) else []
    meta = payload.get("meta", {}) if isinstance(payload, dict) else {}
    return [
        _expect("preview_row_count", isinstance(rows, list) and len(rows) == 1),
        _expect("preview_first_row", isinstance(rows, list) and bool(rows) and rows[0].get("姓名") == "李四"),
        _expect("preview_meta_truncated", isinstance(meta, dict) and meta.get("truncated") is True),
        _expect("preview_meta_max_cells", isinstance(meta, dict) and meta.get("max_cells") == 20),
        _expect("preview_meta_offset_rows", isinstance(meta, dict) and meta.get("offset_rows") == 1),
        _expect("preview_meta_limit_rows", isinstance(meta, dict) and meta.get("limit_rows") == 1),
    ]


def _large_excel_paging_evidence(path: Path) -> list[dict[str, Any]]:
    payload = _read_json(path)
    page = payload.get("page", []) if isinstance(payload, dict) else []
    page_meta = payload.get("page_meta", {}) if isinstance(payload, dict) else {}
    tail = payload.get("tail", []) if isinstance(payload, dict) else []
    tail_meta = payload.get("tail_meta", {}) if isinstance(payload, dict) else {}
    return [
        _expect("page_row_count", isinstance(page, list) and len(page) == 25),
        _expect("page_first_row", isinstance(page, list) and bool(page) and page[0].get("序号") == 1001),
        _expect("page_last_row", isinstance(page, list) and bool(page) and page[-1].get("序号") == 1025),
        _expect("page_meta_offset", isinstance(page_meta, dict) and page_meta.get("offset_rows") == 1000),
        _expect("page_meta_limit", isinstance(page_meta, dict) and page_meta.get("limit_rows") == 25),
        _expect("page_meta_max_cells", isinstance(page_meta, dict) and page_meta.get("max_cells") == 208),
        _expect("tail_row_count", isinstance(tail, list) and len(tail) == 10),
        _expect("tail_first_row", isinstance(tail, list) and bool(tail) and tail[0].get("序号") == 11991),
        _expect("tail_last_row", isinstance(tail, list) and bool(tail) and tail[-1].get("序号") == 12000),
        _expect("tail_meta_offset", isinstance(tail_meta, dict) and tail_meta.get("offset_rows") == 11990),
        _expect("tail_meta_truncated", isinstance(tail_meta, dict) and tail_meta.get("truncated") is True),
    ]


def _worksheet_has_conditional_range(worksheet: Any, range_text: str) -> bool:
    expected = range_text.upper()
    return any(expected in str(item.sqref).replace("$", "").upper().split() for item in worksheet.conditional_formatting)


def _write_excel_evidence(project_root: Path, started_at: float) -> list[dict[str, Any]]:
    rows_path = project_root / "test-plans/basic/write-excel/output/json/loaded-rows.json"
    workbook_path = project_root / "test-plans/basic/write-excel/output/excel/accounts.xlsx"
    payload = _read_json(rows_path)
    rows = payload.get("rows", []) if isinstance(payload, dict) else []
    meta = payload.get("meta", {}) if isinstance(payload, dict) else {}
    return [
        _expect("workbook_fresh", _file_nonempty_after(workbook_path, started_at)),
        _expect("json_fresh", _file_nonempty_after(rows_path, started_at)),
        _expect("row_count", isinstance(rows, list) and len(rows) == 2),
        _expect("meta_sheet", isinstance(meta, dict) and meta.get("sheet") == "名单"),
    ]


def _load_excel_evidence(project_root: Path, started_at: float) -> list[dict[str, Any]]:
    rows_path = project_root / "test-plans/data-driven/load-excel/output/json/finance-people.json"
    workbook_path = project_root / "test-plans/data-driven/load-excel/output/excel/财务在职人员.xlsx"
    payload = _read_json(rows_path)
    rows = payload.get("rows", []) if isinstance(payload, dict) else []
    meta = payload.get("meta", {}) if isinstance(payload, dict) else {}
    return [
        _expect("json_fresh", _file_nonempty_after(rows_path, started_at)),
        _expect("workbook_fresh", _file_nonempty_after(workbook_path, started_at)),
        _expect("finance_row_count", isinstance(rows, list) and len(rows) == 2),
        _expect("meta_sheet", isinstance(meta, dict) and meta.get("sheet") == "名单"),
    ]


def _table_transform_evidence(project_root: Path, started_at: float) -> list[dict[str, Any]]:
    payload_path = project_root / "test-plans/data-driven/table-transform/output/json/table-result.json"
    workbook_path = project_root / "test-plans/data-driven/table-transform/output/excel/table-transform.xlsx"
    payload = _read_json(payload_path)
    rows = payload.get("rows", []) if isinstance(payload, dict) else []
    summary = payload.get("summary", []) if isinstance(payload, dict) else []
    workbook_sheets = _workbook_sheetnames(workbook_path) if workbook_path.exists() else []
    return [
        _expect("json_fresh", _file_nonempty_after(payload_path, started_at)),
        _expect("workbook_fresh", _file_nonempty_after(workbook_path, started_at)),
        _expect("joined_row_count", isinstance(rows, list) and len(rows) == 1),
        _expect("joined_label", isinstance(rows, list) and bool(rows) and rows[0].get("标签") == "财务-A001-王经理"),
        _expect("tax_included_amount", isinstance(rows, list) and bool(rows) and rows[0].get("含税金额") == 1590),
        _expect("summary_contains_finance", _contains_row(summary, {"部门": "财务", "笔数": 2, "总金额": 2700, "平均金额": 1350})),
        _expect("multi_sheet_workbook", {"明细", "部门汇总"}.issubset(set(workbook_sheets))),
    ]


def _table_cleaning_evidence(project_root: Path, started_at: float) -> list[dict[str, Any]]:
    payload_path = project_root / "test-plans/data-driven/table-cleaning/output/json/table-cleaning.json"
    payload = _read_json(payload_path)
    rows = payload.get("rows", []) if isinstance(payload, dict) else []
    return [
        _expect("json_fresh", _file_nonempty_after(payload_path, started_at)),
        _expect("cleaned_row_count", isinstance(rows, list) and len(rows) == 4),
        _expect(
            "split_replace_lookup_first_row",
            _contains_row(
                rows,
                {
                    "姓名": "张三",
                    "部门短名": "财务",
                    "状态": "在职",
                    "入职日期": "2026-01-05",
                    "部门全称": "财务部",
                    "联系电话": "86-13800138000",
                    "模糊部门编码": "FIN",
                },
            ),
        ),
        _expect("global_replace_blank_phone", _contains_row(rows, {"姓名": "王五", "手机号": "未填写", "联系电话": "86-未填写"})),
        _expect("date_parse_slash_dash_dot", _contains_row(rows, {"姓名": "李四", "入职日期": "2026-02-10"})),
        _expect(
            "normalize_headers_union_row",
            _contains_row(rows, {"姓名": "赵六", "部门编码": "FIN", "入职日期": "2026/04/01", "模糊部门编码": "FIN"}),
        ),
        _expect("fuzzy_lookup_score", _row_number_at_least(rows, {"姓名": "张三"}, "部门匹配分", 0.99)),
    ]


def _enterprise_cookbook_evidence(project_root: Path, started_at: float) -> list[dict[str, Any]]:
    root = project_root / "test-plans/data-driven/enterprise-cookbook"
    summary_path = root / "output/json/enterprise-summary.json"
    exported_path = root / "output/json/enterprise-exported.json"
    workbook_path = root / "output/excel/企业数据处理结果.xlsx"
    payload = _read_json(summary_path)
    exported_payload = _read_json(exported_path)
    employees = payload.get("employees", []) if isinstance(payload, dict) else []
    salary_pivot = payload.get("salary_pivot", []) if isinstance(payload, dict) else []
    finance_summary = payload.get("finance_summary", []) if isinstance(payload, dict) else []
    meta = payload.get("meta", {}) if isinstance(payload, dict) else {}
    exported_workbook = exported_payload.get("workbook", {}) if isinstance(exported_payload, dict) else {}
    workbook_sheets = _workbook_sheetnames(workbook_path) if workbook_path.exists() else []
    return [
        _expect("summary_json_fresh", _file_nonempty_after(summary_path, started_at)),
        _expect("exported_json_fresh", _file_nonempty_after(exported_path, started_at)),
        _expect("workbook_fresh", _file_nonempty_after(workbook_path, started_at)),
        _expect("active_employee_count", isinstance(employees, list) and len(employees) == 3),
        _expect("finance_bonus_filled", _contains_row(employees, {"工号": "E001", "奖金": 1000, "部门名称": "财务部"})),
        _expect("hr_blank_bonus_filled", _contains_row(employees, {"工号": "E002", "奖金": 0, "部门名称": "人事部"})),
        _expect("salary_pivot_finance", _contains_row(salary_pivot, {"部门名称": "财务部", "P2": 12000, "P1": 0})),
        _expect("salary_pivot_hr", _contains_row(salary_pivot, {"部门名称": "人事部", "P2": 0, "P1": 9000})),
        _expect(
            "finance_summary_finance",
            _contains_row(
                finance_summary,
                {"部门编码": "FIN", "流水笔数": 2, "流水金额": 3200, "税额合计": 192, "部门名称": "财务部"},
            ),
        ),
        _expect("multi_read_aliases", set(meta.get("value_names", [])) == {"employees", "departments", "transactions"}),
        _expect("multi_sheet_workbook", {"在职员工", "薪资透视", "财务汇总"}.issubset(set(workbook_sheets))),
        _expect("start_cell_header", _excel_cell_value(workbook_path, "在职员工", "B3") == "工号"),
        _expect("exported_active_rows", isinstance(exported_workbook.get("active_export"), list) and len(exported_workbook["active_export"]) == 3),
    ]


def _workbook_sheetnames(path: Path) -> list[str]:
    from openpyxl import load_workbook

    workbook = load_workbook(path, read_only=True)
    try:
        return list(workbook.sheetnames)
    finally:
        workbook.close()


def _excel_cell_value(path: Path, sheet_name: str, address: str) -> Any:
    from openpyxl import load_workbook

    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        return workbook[sheet_name][address].value
    finally:
        workbook.close()


def _contains_row(rows: Any, expected: dict[str, Any]) -> bool:
    if not isinstance(rows, list):
        return False
    for row in rows:
        if isinstance(row, dict) and all(row.get(key) == value for key, value in expected.items()):
            return True
    return False


def _row_number_at_least(rows: Any, expected: dict[str, Any], column: str, minimum: float) -> bool:
    if not isinstance(rows, list):
        return False
    for row in rows:
        if not isinstance(row, dict) or not all(row.get(key) == value for key, value in expected.items()):
            continue
        value = row.get(column)
        return isinstance(value, (int, float)) and value >= minimum
    return False


def _validation_failed_case(name: str, plan_path: Path, validation: Any) -> dict[str, Any]:
    return {
        "name": name,
        "ok": False,
        "plan_path": str(plan_path),
        "validation_ok": False,
        "errors": [error.format() for error in validation.errors],
    }


def _run_failed_case(name: str, plan_path: Path, error: Exception) -> dict[str, Any]:
    return {
        "name": name,
        "ok": False,
        "plan_path": str(plan_path),
        "validation_ok": True,
        "run_ok": False,
        "error": str(error),
        "error_type": type(error).__name__,
    }


def _disable_run_log_echo(_output_dir: Path, logger: Any) -> None:
    logger.echo = False


def _read_json(path: Path) -> Any:
    with path.open("r", encoding="utf-8") as file:
        return json.load(file)


def _file_nonempty(path: Path) -> bool:
    return path.exists() and path.is_file() and path.stat().st_size > 0


def _file_nonempty_after(path: Path, started_at: float) -> bool:
    return _file_nonempty(path) and path.stat().st_mtime >= started_at - 1.0


def _expect(name: str, ok: bool) -> dict[str, Any]:
    return {"name": name, "ok": bool(ok)}


def _cplan_script_path() -> str:
    return ".\\cplan.py" if platform.system() == "Windows" else "./cplan.py"
